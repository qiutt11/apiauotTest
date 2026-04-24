"""Excel 驱动用例的单元测试。

覆盖范围：
    - load_excel_rows()：基础加载、JSON 解析、空行跳过、空文件处理
    - _flatten_excel_validations()：简单值、嵌套 dict、list、深层嵌套
    - _build_body_from_excel()：直接映射（true）、field_mapping 映射
    - _build_excel_validations()：prefix + field_mapping 组合
"""

import json
import os

import pytest


FIXTURES_DIR = os.path.join(os.path.dirname(__file__), "fixtures")


# ==========================================================================
# load_excel_rows() 测试
# ==========================================================================
class TestLoadExcelRows:
    """测试 Excel 数据行加载函数。"""

    def test_basic_load(self, tmp_path):
        """基础场景：加载简单的字符串和数字。"""
        from openpyxl import Workbook
        from common.data_loader import load_excel_rows

        wb = Workbook()
        ws = wb.active
        ws.append(["name", "age", "phone"])
        ws.append(["张三", 25, "13800001111"])
        ws.append(["李四", 30, "13900002222"])
        xlsx = str(tmp_path / "test.xlsx")
        wb.save(xlsx)

        rows = load_excel_rows(xlsx)
        assert len(rows) == 2
        assert rows[0] == {"name": "张三", "age": 25, "phone": "13800001111"}
        assert rows[1] == {"name": "李四", "age": 30, "phone": "13900002222"}

    def test_json_parsing_list(self, tmp_path):
        """JSON 字符串中的 list 应被自动解析为 Python list。"""
        from openpyxl import Workbook
        from common.data_loader import load_excel_rows

        wb = Workbook()
        ws = wb.active
        ws.append(["name", "tags"])
        ws.append(["张三", '["vip", "new"]'])
        xlsx = str(tmp_path / "test.xlsx")
        wb.save(xlsx)

        rows = load_excel_rows(xlsx)
        assert rows[0]["tags"] == ["vip", "new"]

    def test_json_parsing_dict(self, tmp_path):
        """JSON 字符串中的 dict 应被自动解析为 Python dict。"""
        from openpyxl import Workbook
        from common.data_loader import load_excel_rows

        wb = Workbook()
        ws = wb.active
        ws.append(["name", "address"])
        ws.append(["张三", '{"city": "北京", "street": "xx路"}'])
        xlsx = str(tmp_path / "test.xlsx")
        wb.save(xlsx)

        rows = load_excel_rows(xlsx)
        assert rows[0]["address"] == {"city": "北京", "street": "xx路"}

    def test_json_string_not_parsed(self, tmp_path):
        """JSON 能解析为简单值（如纯数字字符串 "123"）时不转换，保留原字符串。"""
        from openpyxl import Workbook
        from common.data_loader import load_excel_rows

        wb = Workbook()
        ws = wb.active
        ws.append(["name", "code"])
        ws.append(["张三", "ABC123"])  # 非 JSON，保留原值
        xlsx = str(tmp_path / "test.xlsx")
        wb.save(xlsx)

        rows = load_excel_rows(xlsx)
        assert rows[0]["code"] == "ABC123"

    def test_skip_empty_rows(self, tmp_path):
        """空行（所有单元格为空）应被跳过。"""
        from openpyxl import Workbook
        from common.data_loader import load_excel_rows

        wb = Workbook()
        ws = wb.active
        ws.append(["name", "age"])
        ws.append(["张三", 25])
        ws.append([None, None])  # 空行
        ws.append(["李四", 30])
        xlsx = str(tmp_path / "test.xlsx")
        wb.save(xlsx)

        rows = load_excel_rows(xlsx)
        assert len(rows) == 2
        assert rows[0]["name"] == "张三"
        assert rows[1]["name"] == "李四"

    def test_skip_empty_header(self, tmp_path):
        """空列头（None）的列应被忽略。"""
        from openpyxl import Workbook
        from common.data_loader import load_excel_rows

        wb = Workbook()
        ws = wb.active
        ws.append(["name", None, "age"])
        ws.append(["张三", "ignored", 25])
        xlsx = str(tmp_path / "test.xlsx")
        wb.save(xlsx)

        rows = load_excel_rows(xlsx)
        assert "name" in rows[0]
        assert "age" in rows[0]
        assert len(rows[0]) == 2  # 空列头的列被跳过

    def test_empty_file(self, tmp_path):
        """只有表头没有数据行时返回空列表。"""
        from openpyxl import Workbook
        from common.data_loader import load_excel_rows

        wb = Workbook()
        ws = wb.active
        ws.append(["name", "age"])
        xlsx = str(tmp_path / "test.xlsx")
        wb.save(xlsx)

        rows = load_excel_rows(xlsx)
        assert rows == []

    def test_completely_empty_file(self, tmp_path):
        """完全空的 Excel（连表头都没有）返回空列表。"""
        from openpyxl import Workbook
        from common.data_loader import load_excel_rows

        wb = Workbook()
        xlsx = str(tmp_path / "test.xlsx")
        wb.save(xlsx)

        rows = load_excel_rows(xlsx)
        assert rows == []

    def test_partial_empty_cells(self, tmp_path):
        """部分单元格为空时，对应字段不出现在字典中。"""
        from openpyxl import Workbook
        from common.data_loader import load_excel_rows

        wb = Workbook()
        ws = wb.active
        ws.append(["name", "age", "phone"])
        ws.append(["张三", None, "13800001111"])  # age 为空
        xlsx = str(tmp_path / "test.xlsx")
        wb.save(xlsx)

        rows = load_excel_rows(xlsx)
        assert len(rows) == 1
        assert "name" in rows[0]
        assert "age" not in rows[0]
        assert "phone" in rows[0]

    def test_nested_json_in_list(self, tmp_path):
        """数组中嵌套对象的 JSON 应正确解析。"""
        from openpyxl import Workbook
        from common.data_loader import load_excel_rows

        wb = Workbook()
        ws = wb.active
        ws.append(["name", "items"])
        ws.append(["张三", '[{"id": 1, "name": "item1"}, {"id": 2, "name": "item2"}]'])
        xlsx = str(tmp_path / "test.xlsx")
        wb.save(xlsx)

        rows = load_excel_rows(xlsx)
        assert rows[0]["items"] == [{"id": 1, "name": "item1"}, {"id": 2, "name": "item2"}]


# ==========================================================================
# _flatten_excel_validations() 测试
# ==========================================================================
class TestFlattenExcelValidations:
    """测试递归展开断言的函数。"""

    def test_simple_string(self):
        """简单字符串值 → 单条 eq 断言。"""
        from conftest import _flatten_excel_validations

        result = _flatten_excel_validations("$.data.name", "张三")
        assert result == [{"eq": ["$.data.name", "张三"]}]

    def test_simple_int(self):
        """整数值 → 单条 eq 断言。"""
        from conftest import _flatten_excel_validations

        result = _flatten_excel_validations("$.data.age", 25)
        assert result == [{"eq": ["$.data.age", 25]}]

    def test_simple_bool(self):
        """布尔值 → 单条 eq 断言。"""
        from conftest import _flatten_excel_validations

        result = _flatten_excel_validations("$.data.active", True)
        assert result == [{"eq": ["$.data.active", True]}]

    def test_dict_flat(self):
        """平铺 dict → 展开为多条带 .key 后缀的断言。"""
        from conftest import _flatten_excel_validations

        result = _flatten_excel_validations("$.data.address", {"city": "北京", "street": "xx路"})
        assert {"eq": ["$.data.address.city", "北京"]} in result
        assert {"eq": ["$.data.address.street", "xx路"]} in result
        assert len(result) == 2

    def test_list_flat(self):
        """list → 展开为带 [index] 后缀的断言。"""
        from conftest import _flatten_excel_validations

        result = _flatten_excel_validations("$.data.tags", ["vip", "new"])
        assert result == [
            {"eq": ["$.data.tags[0]", "vip"]},
            {"eq": ["$.data.tags[1]", "new"]},
        ]

    def test_nested_dict_in_dict(self):
        """嵌套 dict → 递归展开为 $.data.a.b.c 形式。"""
        from conftest import _flatten_excel_validations

        result = _flatten_excel_validations("$.data", {"user": {"name": "张三", "age": 25}})
        assert {"eq": ["$.data.user.name", "张三"]} in result
        assert {"eq": ["$.data.user.age", 25]} in result

    def test_list_of_dicts(self):
        """数组内嵌套对象 → 递归展开为 $.data.items[0].name 形式。"""
        from conftest import _flatten_excel_validations

        result = _flatten_excel_validations("$.data.items", [
            {"name": "item1", "price": 10},
            {"name": "item2", "price": 20},
        ])
        assert {"eq": ["$.data.items[0].name", "item1"]} in result
        assert {"eq": ["$.data.items[0].price", 10]} in result
        assert {"eq": ["$.data.items[1].name", "item2"]} in result
        assert {"eq": ["$.data.items[1].price", 20]} in result
        assert len(result) == 4

    def test_empty_dict(self):
        """空 dict → 无断言。"""
        from conftest import _flatten_excel_validations

        result = _flatten_excel_validations("$.data", {})
        assert result == []

    def test_empty_list(self):
        """空 list → 无断言。"""
        from conftest import _flatten_excel_validations

        result = _flatten_excel_validations("$.data.tags", [])
        assert result == []


# ==========================================================================
# _build_body_from_excel() 测试
# ==========================================================================
class TestBuildBodyFromExcel:
    """测试 Excel 行数据转请求 body。"""

    def test_true_direct_mapping(self):
        """body_from_excel: true → Excel 列名直接作为字段名。"""
        from conftest import _build_body_from_excel

        row = {"name": "张三", "age": 25, "tags": ["vip"]}
        body = _build_body_from_excel(row, True)
        assert body == {"name": "张三", "age": 25, "tags": ["vip"]}

    def test_true_returns_copy(self):
        """返回的应是副本，不影响原始 row_data。"""
        from conftest import _build_body_from_excel

        row = {"name": "张三"}
        body = _build_body_from_excel(row, True)
        body["name"] = "李四"
        assert row["name"] == "张三"  # 原数据未被修改

    def test_field_mapping(self):
        """field_mapping 映射 → Excel 列名转为接口字段名。"""
        from conftest import _build_body_from_excel

        row = {"name": "张三", "phone": "138xxx", "age": 25}
        mapping_config = {"field_mapping": {"name": "userName", "phone": "phoneNumber"}}
        body = _build_body_from_excel(row, mapping_config)
        # 映射的字段
        assert body["userName"] == "张三"
        assert body["phoneNumber"] == "138xxx"
        # 未映射的字段保留原名
        assert body["age"] == 25
        assert "name" not in body
        assert "phone" not in body

    def test_empty_field_mapping(self):
        """空 field_mapping → 等同于 true，保留原列名。"""
        from conftest import _build_body_from_excel

        row = {"name": "张三", "age": 25}
        body = _build_body_from_excel(row, {"field_mapping": {}})
        assert body == {"name": "张三", "age": 25}

    def test_field_mapping_without_key(self):
        """dict 但没有 field_mapping key → 等同于空 mapping。"""
        from conftest import _build_body_from_excel

        row = {"name": "张三"}
        body = _build_body_from_excel(row, {})
        assert body == {"name": "张三"}


# ==========================================================================
# _build_excel_validations() 测试
# ==========================================================================
class TestBuildExcelValidations:
    """测试 Excel 行数据转断言列表。"""

    def test_basic_with_prefix(self):
        """基础场景：prefix + 简单值 → eq 断言。"""
        from conftest import _build_excel_validations

        row = {"name": "张三", "age": 25}
        result = _build_excel_validations(row, {"prefix": "$.data"})
        assert {"eq": ["$.data.name", "张三"]} in result
        assert {"eq": ["$.data.age", 25]} in result

    def test_default_prefix(self):
        """不指定 prefix 时默认为 $.data。"""
        from conftest import _build_excel_validations

        row = {"name": "张三"}
        result = _build_excel_validations(row, {})
        assert result == [{"eq": ["$.data.name", "张三"]}]

    def test_with_field_mapping(self):
        """field_mapping 映射 → Excel 列名转为响应字段名。"""
        from conftest import _build_excel_validations

        row = {"name": "张三", "phone": "138xxx"}
        config = {
            "prefix": "$.data",
            "field_mapping": {"name": "user_name", "phone": "phone_number"},
        }
        result = _build_excel_validations(row, config)
        assert {"eq": ["$.data.user_name", "张三"]} in result
        assert {"eq": ["$.data.phone_number", "138xxx"]} in result

    def test_nested_value_with_mapping(self):
        """嵌套值 + mapping → 递归展开并使用映射后的字段名。"""
        from conftest import _build_excel_validations

        row = {"addr": {"city": "北京", "street": "xx路"}}
        config = {"prefix": "$.data", "field_mapping": {"addr": "address"}}
        result = _build_excel_validations(row, config)
        assert {"eq": ["$.data.address.city", "北京"]} in result
        assert {"eq": ["$.data.address.street", "xx路"]} in result

    def test_complex_row_data(self):
        """复杂场景：混合简单值、数组、嵌套对象。"""
        from conftest import _build_excel_validations

        row = {
            "name": "张三",
            "age": 25,
            "tags": ["vip", "new"],
            "address": {"city": "北京", "street": "xx路"},
        }
        result = _build_excel_validations(row, {"prefix": "$.data"})
        assert {"eq": ["$.data.name", "张三"]} in result
        assert {"eq": ["$.data.age", 25]} in result
        assert {"eq": ["$.data.tags[0]", "vip"]} in result
        assert {"eq": ["$.data.tags[1]", "new"]} in result
        assert {"eq": ["$.data.address.city", "北京"]} in result
        assert {"eq": ["$.data.address.street", "xx路"]} in result
        assert len(result) == 6

"""用例数据加载模块。

支持从三种格式加载测试用例数据：
    - YAML（推荐）：可读性强，支持注释
    - JSON：与接口数据格式一致
    - Excel：适合业务人员批量管理

所有格式统一返回：
    {"module": "模块名", "testcases": [用例1, 用例2, ...]}
"""

import json
import os
from typing import Any

import yaml
from openpyxl import load_workbook


def load_testcases(file_path: str) -> dict[str, Any]:
    """根据文件扩展名自动选择加载器，返回标准化的用例数据。

    Args:
        file_path: 用例文件路径（.yaml/.yml/.json/.xlsx）

    Returns:
        {"module": "模块名", "testcases": [用例字典列表]}

    Raises:
        ValueError: 不支持的文件格式
    """
    ext = os.path.splitext(file_path)[1].lower()
    if ext in (".yaml", ".yml"):
        return _load_yaml(file_path)
    elif ext == ".json":
        return _load_json(file_path)
    elif ext == ".xlsx":
        return _load_excel(file_path)
    else:
        raise ValueError(f"Unsupported file format: {ext}")


def _load_yaml(file_path: str) -> dict:
    """加载 YAML 格式的用例文件。"""
    with open(file_path, "r", encoding="utf-8") as f:
        return yaml.safe_load(f) or {}  # 空文件返回 {} 而非 None


def _load_json(file_path: str) -> dict:
    """加载 JSON 格式的用例文件。"""
    with open(file_path, "r", encoding="utf-8") as f:
        return json.load(f)


def _load_excel(file_path: str) -> dict:
    """加载 Excel 格式的用例文件。

    规则：
        - Sheet 名称作为模块名（module）
        - 第一行为表头（name, method, url, headers, body, extract, validate 等）
        - 每行一个用例
        - headers/body/extract/validate 等字段需要填写 JSON 字符串
        - 空列头（None）自动跳过
        - 完全空的行自动跳过
    """
    wb = load_workbook(file_path, read_only=True)
    try:
        ws = wb.active
        module_name = ws.title  # Sheet 名作为模块名

        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 2:
            return {"module": module_name, "testcases": []}

        # 解析表头（跳过 None 列头）
        headers = [str(h).strip() if h is not None else "" for h in rows[0]]
        testcases = []

        # 需要 JSON 解析的字段
        json_fields = ("headers", "body", "extract", "validate",
                       "db_setup", "db_extract", "db_teardown")

        for row in rows[1:]:
            case = {}
            for i, header in enumerate(headers):
                if not header:  # 跳过空列头
                    continue
                value = row[i] if i < len(row) else None
                if value is None or (isinstance(value, str) and value.strip() == ""):
                    continue
                # JSON 字段需要解析字符串为 dict/list
                if header in json_fields:
                    case[header] = json.loads(str(value))
                else:
                    case[header] = value
            if case:  # 跳过完全空的行
                testcases.append(case)

        return {"module": module_name, "testcases": testcases}
    finally:
        wb.close()  # 确保文件句柄释放（即使解析出错）


def load_excel_rows(file_path: str) -> list[dict]:
    """读取 Excel 每行数据为字典列表（表头作为 key）。

    用于 Excel 驱动的数据参数化场景，每行对应一组测试数据。

    规则：
        - 第一行为表头（字段名）
        - 每行一组数据，跳过空行
        - 单元格值尝试 JSON 解析：能解析为 dict/list 的自动转换，否则保留原值
        - 这样支持复杂结构：数组字段填 ["vip","new"]，嵌套对象填 {"city":"北京"}

    Args:
        file_path: Excel 文件路径（.xlsx）

    Returns:
        字典列表，每个字典代表一行数据 {列名: 值}
    """
    wb = load_workbook(file_path, read_only=True)
    try:
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 2:
            return []

        headers = [str(h).strip() if h is not None else "" for h in rows[0]]
        result = []

        for row in rows[1:]:
            record = {}
            all_empty = True
            for i, header in enumerate(headers):
                if not header:
                    continue
                value = row[i] if i < len(row) else None
                if value is None or (isinstance(value, str) and value.strip() == ""):
                    continue
                all_empty = False
                # openpyxl 将整数单元格返回为 float（如 25 → 25.0），转回 int
                if isinstance(value, float) and value == int(value):
                    value = int(value)
                # 尝试 JSON 解析：dict/list 自动转换，其他保留原值
                if isinstance(value, str):
                    try:
                        parsed = json.loads(value)
                        if isinstance(parsed, (dict, list)):
                            value = parsed
                    except (json.JSONDecodeError, ValueError):
                        pass
                record[header] = value
            if not all_empty:
                result.append(record)

        return result
    finally:
        wb.close()


def scan_testcase_files(directory: str) -> list[str]:
    """扫描目录下所有支持的用例文件，按文件名排序返回。

    Args:
        directory: 要扫描的目录路径

    Returns:
        文件绝对路径列表
    """
    valid_extensions = {".yaml", ".yml", ".json", ".xlsx"}
    files = []
    for root, _, filenames in os.walk(directory):
        for filename in sorted(filenames):
            ext = os.path.splitext(filename)[1].lower()
            if ext in valid_extensions:
                files.append(os.path.join(root, filename))
    return files
